"""Excel parser — sheet-by-sheet extraction with table preservation."""

from __future__ import annotations

import asyncio
from pathlib import Path

from core.models import DocumentType

from .base import BaseParser, ParsedDocument


class ExcelParser(BaseParser):
    """Parse Excel files (.xlsx, .xls) with sheet-level structure."""

    document_type = DocumentType.EXCEL
    supported_mimes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]

    async def parse(self, file_path: Path) -> ParsedDocument:
        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, self._parse_sync, file_path)

    def _parse_sync(self, file_path: Path) -> ParsedDocument:
        import openpyxl

        wb = openpyxl.load_workbook(str(file_path), data_only=True)

        text_parts = []
        tables = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_text = [f"## Sheet: {sheet_name}"]

            # Extract all rows as text
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if any(cell is not None for cell in row):
                    row_text = " | ".join(
                        str(cell) if cell is not None else ""
                        for cell in row
                    )
                    sheet_text.append(row_text)

                    # First row = headers
                    if row_idx == 1:
                        headers = [
                            str(c) if c is not None else ""
                            for c in row
                        ]

            text_parts.append("\n".join(sheet_text))
            tables.append({
                "sheet": sheet_name,
                "rows": ws.max_row - 1,  # minus header
                "columns": ws.max_column,
                "headers": headers if ws.max_row > 0 else [],
            })

        text = "\n\n".join(text_parts)
        metadata = {
            "sheet_count": len(wb.sheetnames),
            "sheets": wb.sheetnames,
            "total_rows": sum(t["rows"] for t in tables),
        }

        return ParsedDocument(
            text=text,
            metadata=metadata,
            tables=tables,
        )

    def can_handle(self, file_path: Path, mime_type: str) -> bool:
        if mime_type in self.supported_mimes:
            return True
        return file_path.suffix.lower() in (".xlsx", ".xls", ".xlsm")
